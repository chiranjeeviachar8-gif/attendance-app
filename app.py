"""
Student Attendance Web Application
-----------------------------------
A Flask web app where teachers register/login and each teacher manages
their OWN private list of students, attendance, and CSV exports.

DATABASE:
- If a DATABASE_URL environment variable is set (e.g. a free Neon/Render
  Postgres connection string), that permanent database is used -- your
  logins and attendance data will NEVER be wiped on restart/redeploy.
- If DATABASE_URL is not set, it falls back to a local SQLite file
  (attendance.db) -- fine for running on your own computer, but NOT
  safe on Render's free tier (gets wiped on restart).

Run locally:
    pip install -r requirements.txt
    python app.py

Then open in your browser: http://127.0.0.1:5000
"""

import os
import sys
import csv
import io
import base64
from datetime import date, timedelta, datetime

from flask import Flask, render_template, request, redirect, url_for, session, flash, Response
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError, OperationalError
from openpyxl import load_workbook
from PIL import Image

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key")
app.permanent_session_lifetime = timedelta(days=3650)  # "Remember Me" ~= stays logged in for 10 years (effectively forever)

# ---------------------------------------------------------------------
# Database setup (works with SQLite locally, PostgreSQL or MySQL when deployed)
# ---------------------------------------------------------------------
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///attendance.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
# Plain "mysql://" needs a driver name for SQLAlchemy -- default to pymysql
if DATABASE_URL.startswith("mysql://"):
    DATABASE_URL = DATABASE_URL.replace("mysql://", "mysql+pymysql://", 1)

if "DATABASE_URL" in os.environ:
    safe_url = DATABASE_URL.split("@")[-1]
    print(f"[DB] DATABASE_URL is SET. Connecting to host: {safe_url}", flush=True)
else:
    print("[DB] WARNING: DATABASE_URL is NOT set in the environment.", flush=True)
    print("[DB] Falling back to local SQLite file (attendance.db).", flush=True)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
IS_POSTGRES = engine.dialect.name == "postgresql"
IS_MYSQL = engine.dialect.name in ("mysql", "mariadb")


def init_db():
    if IS_POSTGRES:
        id_type = "SERIAL PRIMARY KEY"
    elif IS_MYSQL:
        id_type = "INTEGER PRIMARY KEY AUTO_INCREMENT"
    else:
        id_type = "INTEGER PRIMARY KEY AUTOINCREMENT"
    try:
        with engine.begin() as conn:
            conn.execute(text(f"""
                CREATE TABLE IF NOT EXISTS teachers (
                    id {id_type},
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL
                )
            """))
            conn.execute(text(f"""
                CREATE TABLE IF NOT EXISTS students (
                    id {id_type},
                    owner TEXT NOT NULL,
                    roll_no TEXT NOT NULL,
                    name TEXT NOT NULL,
                    register_no TEXT,
                    class_name TEXT,
                    batch_name TEXT,
                    UNIQUE(owner, roll_no)
                )
            """))
            conn.execute(text(f"""
                CREATE TABLE IF NOT EXISTS attendance (
                    id {id_type},
                    student_id INTEGER NOT NULL REFERENCES students(id),
                    date TEXT NOT NULL,
                    status TEXT NOT NULL CHECK(status IN ('Present', 'Absent')),
                    marked_by TEXT,
                    UNIQUE(student_id, date)
                )
            """))
            conn.execute(text(f"""
                CREATE TABLE IF NOT EXISTS holidays (
                    id {id_type},
                    owner TEXT NOT NULL,
                    holiday_date TEXT NOT NULL,
                    description TEXT,
                    UNIQUE(owner, holiday_date)
                )
            """))
        db_label = "Postgres/Neon" if IS_POSTGRES else ("MySQL" if IS_MYSQL else "local SQLite")
        print(f"[DB] SUCCESS: connected and tables ready ({db_label}).", flush=True)
    except OperationalError as e:
        print(f"[DB] FAILED TO CONNECT: {e}", flush=True)
        raise

    # --- Migration: add any columns that older deployments of this DB might be missing ---
    # Runs in its OWN transaction so it can never break the main table setup above.
    def _column_exists(conn, table, column):
        if IS_POSTGRES:
            return conn.execute(text("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = :t AND column_name = :c
            """), {"t": table, "c": column}).first() is not None
        elif IS_MYSQL:
            return conn.execute(text("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name = :t AND column_name = :c AND table_schema = DATABASE()
            """), {"t": table, "c": column}).first() is not None
        else:
            cols = conn.execute(text(f"PRAGMA table_info({table})")).fetchall()
            return any(c[1] == column for c in cols)

    for column in ["owner", "register_no", "batch_name"]:
        try:
            with engine.begin() as conn:
                if not _column_exists(conn, "students", column):
                    conn.execute(text(f"ALTER TABLE students ADD COLUMN {column} TEXT"))
                    print(f"[DB] Migrated: added '{column}' column to students table.", flush=True)
                else:
                    print(f"[DB] '{column}' column already present -- no migration needed.", flush=True)
        except Exception as e:
            print(f"[DB] Migration check for '{column}' failed (non-fatal): {e}", flush=True)

    for column in ["profile_photo", "background_photo"]:
        try:
            with engine.begin() as conn:
                if not _column_exists(conn, "teachers", column):
                    col_type = "LONGTEXT" if IS_MYSQL else "TEXT"
                    conn.execute(text(f"ALTER TABLE teachers ADD COLUMN {column} {col_type}"))
                    print(f"[DB] Migrated: added '{column}' column to teachers table.", flush=True)
                else:
                    print(f"[DB] '{column}' column already present -- no migration needed.", flush=True)
        except Exception as e:
            print(f"[DB] Migration check for '{column}' failed (non-fatal): {e}", flush=True)


def query(sql, params=None, fetch="all"):
    """Run a SQL statement. fetch: 'all' | 'one' | None (for writes)."""
    with engine.begin() as conn:
        result = conn.execute(text(sql), params or {})
        if fetch == "all":
            return result.mappings().all()
        if fetch == "one":
            return result.mappings().first()
        return None


def login_required(func):
    from functools import wraps

    @wraps(func)
    def wrapper(*args, **kwargs):
        if "teacher" not in session:
            flash("Please log in first.", "warning")
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def current_owner():
    """The logged-in teacher's username -- used to scope all their data."""
    return session["teacher"]


def resize_image_to_base64(file_storage, max_size, quality=80):
    """Resize an uploaded image and return it as a base64 JPEG string."""
    img = Image.open(file_storage)
    if img.mode != "RGB":
        img = img.convert("RGB")
    img.thumbnail(max_size, Image.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=quality, optimize=True)
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


@app.context_processor
def inject_teacher_theme():
    """Makes the logged-in teacher's profile photo / background available on every page."""
    if "teacher" in session:
        t = query(
            "SELECT profile_photo, background_photo FROM teachers WHERE username = :u",
            {"u": session["teacher"]}, fetch="one"
        )
        if t:
            return {
                "nav_profile_photo": t["profile_photo"],
                "nav_background_photo": t["background_photo"],
            }
    return {"nav_profile_photo": None, "nav_background_photo": None}


# ---------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------
@app.route("/")
def home():
    if "teacher" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        try:
            query(
                "INSERT INTO teachers (username, password_hash) VALUES (:u, :p)",
                {"u": username, "p": generate_password_hash(password)},
                fetch=None,
            )
            flash("Registration successful! You can now log in.", "success")
            return redirect(url_for("login"))
        except IntegrityError:
            flash("That username is already taken.", "danger")

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        remember = request.form.get("remember") == "on"

        teacher = query(
            "SELECT * FROM teachers WHERE username = :u", {"u": username}, fetch="one"
        )

        if teacher and check_password_hash(teacher["password_hash"], password):
            session.permanent = remember
            session["teacher"] = username
            flash(f"Welcome back, {username}!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("teacher", None)
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ---------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------
@app.route("/dashboard")
@login_required
def dashboard():
    owner = current_owner()
    student_count = query(
        "SELECT COUNT(*) AS c FROM students WHERE owner = :o", {"o": owner}, fetch="one"
    )["c"]
    today = str(date.today())
    today_marked = query("""
        SELECT COUNT(*) AS c FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE a.date = :d AND s.owner = :o
    """, {"d": today, "o": owner}, fetch="one")["c"]
    return render_template(
        "dashboard.html",
        teacher=owner,
        student_count=student_count,
        today=today,
        today_marked=today_marked,
    )


# ---------------------------------------------------------------------
# Student management (each teacher only sees/edits their OWN students)
# ---------------------------------------------------------------------
@app.route("/students", methods=["GET", "POST"])
@login_required
def students():
    owner = current_owner()
    if request.method == "POST":
        roll_no = request.form["roll_no"].strip()
        name = request.form["name"].strip()
        register_no = request.form.get("register_no", "").strip()
        class_name = request.form["class_name"].strip()
        batch_name = request.form.get("batch_name", "").strip()
        try:
            query(
                """INSERT INTO students (owner, roll_no, name, register_no, class_name, batch_name)
                   VALUES (:o, :r, :n, :reg, :c, :b)""",
                {"o": owner, "r": roll_no, "n": name, "reg": register_no, "c": class_name, "b": batch_name},
                fetch=None,
            )
            flash(f"Student '{name}' added.", "success")
        except IntegrityError:
            flash("You already have a student with that roll number.", "danger")

    all_students = query(
        "SELECT * FROM students WHERE owner = :o ORDER BY CAST(roll_no AS INTEGER), roll_no",
        {"o": owner},
    )
    return render_template("students.html", students=all_students)


@app.route("/students/import", methods=["POST"])
@login_required
def import_students():
    owner = current_owner()
    file = request.files.get("excel_file")

    if not file or file.filename == "":
        flash("Please choose an Excel file to upload.", "danger")
        return redirect(url_for("students"))

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Only .xlsx or .xlsm Excel files are supported.", "danger")
        return redirect(url_for("students"))

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"Could not read that Excel file: {e}", "danger")
        return redirect(url_for("students"))

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash("That Excel file appears to be empty.", "danger")
        return redirect(url_for("students"))

    # First row = headers. Match "Roll No", "Name", "Class" (case-insensitive, flexible spacing).
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    def find_col(*names):
        for i, h in enumerate(headers):
            if h in names:
                return i
        return None

    roll_col = find_col("roll no", "roll_no", "rollno", "roll")
    name_col = find_col("name", "student name")
    register_col = find_col("register no", "register_no", "registerno", "reg no", "reg_no")
    class_col = find_col("class", "class_name", "class name")
    batch_col = find_col("batch", "batch_name", "batch name")

    if roll_col is None or name_col is None:
        flash('Excel must have column headers "Roll No" and "Name" in the first row '
              '("Register No", "Class", "Batch" are optional).', "danger")
        return redirect(url_for("students"))

    def cell_value(row, col):
        if col is not None and col < len(row) and row[col] is not None:
            return str(row[col]).strip()
        return ""

    added, skipped = 0, 0
    for row in rows[1:]:
        if row is None or all(cell in (None, "") for cell in row):
            continue
        roll_no = cell_value(row, roll_col)
        name = cell_value(row, name_col)
        register_no = cell_value(row, register_col)
        class_name = cell_value(row, class_col)
        batch_name = cell_value(row, batch_col)

        if not roll_no or not name:
            skipped += 1
            continue

        try:
            query(
                """INSERT INTO students (owner, roll_no, name, register_no, class_name, batch_name)
                   VALUES (:o, :r, :n, :reg, :c, :b)""",
                {"o": owner, "r": roll_no, "n": name, "reg": register_no, "c": class_name, "b": batch_name},
                fetch=None,
            )
            added += 1
        except IntegrityError:
            skipped += 1

    flash(f"Import complete: {added} student(s) added, {skipped} skipped "
          f"(duplicates or missing data).", "success" if added else "warning")
    return redirect(url_for("students"))


@app.route("/students/delete/<int:student_id>", methods=["POST"])
@login_required
def delete_student(student_id):
    owner = current_owner()
    student = query(
        "SELECT * FROM students WHERE id = :id AND owner = :o", {"id": student_id, "o": owner}, fetch="one"
    )
    if student:
        query("DELETE FROM attendance WHERE student_id = :id", {"id": student_id}, fetch=None)
        query("DELETE FROM students WHERE id = :id", {"id": student_id}, fetch=None)
        flash(f"Student '{student['name']}' and their attendance records were deleted.", "info")
    else:
        flash("Student not found.", "danger")
    return redirect(url_for("students"))


@app.route("/students/edit/<int:student_id>", methods=["GET", "POST"])
@login_required
def edit_student(student_id):
    owner = current_owner()
    student = query(
        "SELECT * FROM students WHERE id = :id AND owner = :o", {"id": student_id, "o": owner}, fetch="one"
    )
    if not student:
        flash("Student not found.", "danger")
        return redirect(url_for("students"))

    if request.method == "POST":
        roll_no = request.form["roll_no"].strip()
        name = request.form["name"].strip()
        register_no = request.form.get("register_no", "").strip()
        class_name = request.form["class_name"].strip()
        batch_name = request.form.get("batch_name", "").strip()
        try:
            query(
                """UPDATE students SET roll_no = :r, name = :n, register_no = :reg,
                   class_name = :c, batch_name = :b WHERE id = :id AND owner = :o""",
                {"r": roll_no, "n": name, "reg": register_no, "c": class_name, "b": batch_name,
                 "id": student_id, "o": owner},
                fetch=None,
            )
            flash(f"Student '{name}' updated.", "success")
            return redirect(url_for("students"))
        except IntegrityError:
            flash("You already have another student with that roll number.", "danger")
            student = query(
                "SELECT * FROM students WHERE id = :id AND owner = :o", {"id": student_id, "o": owner}, fetch="one"
            )

    return render_template("edit_student.html", student=student)


# ---------------------------------------------------------------------
# Attendance (scoped to the logged-in teacher's own students)
# ---------------------------------------------------------------------
@app.route("/attendance/mark", methods=["GET", "POST"])
@login_required
def mark_attendance():
    owner = current_owner()
    all_students = query(
        "SELECT * FROM students WHERE owner = :o ORDER BY CAST(roll_no AS INTEGER), roll_no",
        {"o": owner},
    )
    selected_date = request.values.get("att_date") or str(date.today())

    holiday = query(
        "SELECT * FROM holidays WHERE owner = :o AND holiday_date = :d",
        {"o": owner, "d": selected_date}, fetch="one"
    )

    if request.method == "POST":
        if holiday:
            flash(f"{selected_date} is marked as a holiday ({holiday['description'] or 'Holiday'}) — attendance was not saved.", "warning")
        else:
            for student in all_students:
                status = request.form.get(f"status_{student['id']}")
                if status in ("Present", "Absent"):
                    if IS_MYSQL:
                        upsert_sql = """
                            INSERT INTO attendance (student_id, date, status, marked_by)
                            VALUES (:sid, :d, :s, :m)
                            ON DUPLICATE KEY UPDATE status = VALUES(status), marked_by = VALUES(marked_by)
                        """
                    else:
                        upsert_sql = """
                            INSERT INTO attendance (student_id, date, status, marked_by)
                            VALUES (:sid, :d, :s, :m)
                            ON CONFLICT (student_id, date)
                            DO UPDATE SET status = excluded.status, marked_by = excluded.marked_by
                        """
                    query(upsert_sql, {"sid": student["id"], "d": selected_date, "s": status, "m": owner}, fetch=None)
            flash(f"Attendance for {selected_date} saved.", "success")

    existing = query("""
        SELECT a.student_id, a.status FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE a.date = :d AND s.owner = :o
    """, {"d": selected_date, "o": owner})
    existing_map = {row["student_id"]: row["status"] for row in existing}

    return render_template(
        "mark_attendance.html",
        students=all_students,
        selected_date=selected_date,
        existing_map=existing_map,
        holiday=holiday,
    )


@app.route("/attendance/by-date", methods=["GET", "POST"])
@login_required
def view_by_date():
    owner = current_owner()
    selected_date = request.values.get("att_date") or str(date.today())

    rows = query("""
        SELECT s.roll_no, s.name, a.status
        FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE a.date = :d AND s.owner = :o
        ORDER BY CAST(s.roll_no AS INTEGER), s.roll_no
    """, {"d": selected_date, "o": owner})

    present_count = sum(1 for r in rows if r["status"] == "Present")
    return render_template(
        "view_by_date.html",
        rows=rows,
        selected_date=selected_date,
        present_count=present_count,
        total=len(rows),
    )


@app.route("/attendance/by-student", methods=["GET", "POST"])
@login_required
def view_by_student():
    owner = current_owner()
    roll_no = request.values.get("roll_no", "").strip()
    rows = []
    student_name = None
    present_pct = None

    if roll_no:
        student = query(
            "SELECT * FROM students WHERE roll_no = :r AND owner = :o", {"r": roll_no, "o": owner}, fetch="one"
        )
        if student:
            student_name = student["name"]
            rows = query(
                "SELECT date, status FROM attendance WHERE student_id = :sid ORDER BY date",
                {"sid": student["id"]},
            )
            if rows:
                present = sum(1 for r in rows if r["status"] == "Present")
                present_pct = round(present / len(rows) * 100, 1)
        else:
            flash("No student found with that roll number in your class.", "danger")

    return render_template(
        "view_by_student.html",
        rows=rows,
        roll_no=roll_no,
        student_name=student_name,
        present_pct=present_pct,
    )


# ---------------------------------------------------------------------
# CSV Export (only the logged-in teacher's own data, named with their username)
# ---------------------------------------------------------------------
@app.route("/export/csv")
@login_required
def export_csv():
    owner = current_owner()
    rows = query("""
        SELECT s.roll_no, s.name, s.register_no, s.class_name, s.batch_name, a.date, a.status, a.marked_by
        FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE s.owner = :o
        ORDER BY a.date, CAST(s.roll_no AS INTEGER), s.roll_no
    """, {"o": owner})

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Roll No", "Name", "Register No", "Class", "Batch", "Date", "Status", "Marked By"])
    for r in rows:
        writer.writerow([r["roll_no"], r["name"], r["register_no"] or "", r["class_name"] or "",
                          r["batch_name"] or "", r["date"], r["status"], r["marked_by"] or ""])

    filename = f"attendance_{owner}_{date.today()}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/export/csv/date")
@login_required
def export_csv_by_date():
    owner = current_owner()
    selected_date = request.args.get("att_date") or str(date.today())

    rows = query("""
        SELECT s.roll_no, s.name, s.register_no, s.class_name, s.batch_name, a.date, a.status, a.marked_by
        FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE a.date = :d AND s.owner = :o
        ORDER BY CAST(s.roll_no AS INTEGER), s.roll_no
    """, {"d": selected_date, "o": owner})

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Roll No", "Name", "Register No", "Class", "Batch", "Date", "Status", "Marked By"])
    for r in rows:
        writer.writerow([r["roll_no"], r["name"], r["register_no"] or "", r["class_name"] or "",
                          r["batch_name"] or "", r["date"], r["status"], r["marked_by"] or ""])

    filename = f"attendance_{owner}_{selected_date}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------
# Holidays
# ---------------------------------------------------------------------
@app.route("/holidays", methods=["GET", "POST"])
@login_required
def holidays():
    owner = current_owner()
    if request.method == "POST":
        holiday_date = request.form["holiday_date"].strip()
        description = request.form.get("description", "").strip()
        try:
            query(
                "INSERT INTO holidays (owner, holiday_date, description) VALUES (:o, :d, :desc)",
                {"o": owner, "d": holiday_date, "desc": description},
                fetch=None,
            )
            flash(f"Holiday added for {holiday_date}.", "success")
        except IntegrityError:
            flash(f"{holiday_date} is already marked as a holiday.", "danger")

    all_holidays = query(
        "SELECT * FROM holidays WHERE owner = :o ORDER BY holiday_date",
        {"o": owner},
    )
    return render_template("holidays.html", holidays=all_holidays)


@app.route("/holidays/delete/<int:holiday_id>", methods=["POST"])
@login_required
def delete_holiday(holiday_id):
    owner = current_owner()
    query(
        "DELETE FROM holidays WHERE id = :id AND owner = :o",
        {"id": holiday_id, "o": owner}, fetch=None,
    )
    flash("Holiday removed.", "info")
    return redirect(url_for("holidays"))


# ---------------------------------------------------------------------
# Profile (photo, background image, quick stats)
# ---------------------------------------------------------------------
@app.route("/profile")
@login_required
def profile():
    owner = current_owner()
    student_count = query(
        "SELECT COUNT(*) AS c FROM students WHERE owner = :o", {"o": owner}, fetch="one"
    )["c"]
    days_marked = query("""
        SELECT COUNT(DISTINCT a.date) AS c FROM attendance a
        JOIN students s ON s.id = a.student_id WHERE s.owner = :o
    """, {"o": owner}, fetch="one")["c"]
    total_present = query("""
        SELECT COUNT(*) AS c FROM attendance a
        JOIN students s ON s.id = a.student_id WHERE s.owner = :o AND a.status = 'Present'
    """, {"o": owner}, fetch="one")["c"]
    total_absent = query("""
        SELECT COUNT(*) AS c FROM attendance a
        JOIN students s ON s.id = a.student_id WHERE s.owner = :o AND a.status = 'Absent'
    """, {"o": owner}, fetch="one")["c"]

    return render_template(
        "profile.html",
        teacher=owner,
        student_count=student_count,
        days_marked=days_marked,
        total_present=total_present,
        total_absent=total_absent,
    )


@app.route("/profile/photo", methods=["POST"])
@login_required
def upload_profile_photo():
    owner = current_owner()
    file = request.files.get("photo")
    if not file or file.filename == "":
        flash("Please choose an image file.", "danger")
        return redirect(url_for("profile"))
    try:
        photo_b64 = resize_image_to_base64(file, max_size=(300, 300), quality=80)
        query(
            "UPDATE teachers SET profile_photo = :p WHERE username = :u",
            {"p": photo_b64, "u": owner}, fetch=None,
        )
        flash("Profile photo updated.", "success")
    except Exception as e:
        flash(f"Could not process that image: {e}", "danger")
    return redirect(url_for("profile"))


@app.route("/profile/background", methods=["POST"])
@login_required
def upload_background_photo():
    owner = current_owner()
    file = request.files.get("background")
    if not file or file.filename == "":
        flash("Please choose an image file.", "danger")
        return redirect(url_for("profile"))
    try:
        bg_b64 = resize_image_to_base64(file, max_size=(1600, 1600), quality=70)
        query(
            "UPDATE teachers SET background_photo = :p WHERE username = :u",
            {"p": bg_b64, "u": owner}, fetch=None,
        )
        flash("Background image updated.", "success")
    except Exception as e:
        flash(f"Could not process that image: {e}", "danger")
    return redirect(url_for("profile"))


@app.route("/profile/background/remove", methods=["POST"])
@login_required
def remove_background_photo():
    owner = current_owner()
    query(
        "UPDATE teachers SET background_photo = NULL WHERE username = :u",
        {"u": owner}, fetch=None,
    )
    flash("Background image removed.", "info")
    return redirect(url_for("profile"))


# ---------------------------------------------------------------------
# Attendance History (all dates, with day-of-week, drill down by date)
# ---------------------------------------------------------------------
@app.route("/attendance/history")
@login_required
def attendance_history():
    owner = current_owner()
    rows = query("""
        SELECT a.date AS att_date,
               SUM(CASE WHEN a.status = 'Present' THEN 1 ELSE 0 END) AS present_count,
               SUM(CASE WHEN a.status = 'Absent' THEN 1 ELSE 0 END) AS absent_count,
               COUNT(*) AS total
        FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE s.owner = :o
        GROUP BY a.date
        ORDER BY a.date DESC
    """, {"o": owner})

    history = []
    for r in rows:
        try:
            day_name = datetime.strptime(r["att_date"], "%Y-%m-%d").strftime("%A")
        except ValueError:
            day_name = ""
        history.append({
            "date": r["att_date"],
            "day_name": day_name,
            "present": r["present_count"],
            "absent": r["absent_count"],
            "total": r["total"],
        })

    return render_template("attendance_history.html", history=history)


# ---------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------
init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")
